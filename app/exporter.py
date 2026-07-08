"""导出完成记录为 Excel(.xlsx)。文件名包含姓名与日期。"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["序号", "完成时间", "任务模板", "提醒内容", "操作人", "提醒时间", "延迟次数", "响应时长"]


def _response_duration(reminded_at: str, completed_at: str) -> str:
    """计算响应时长（从弹窗到点完成），返回易读字符串。"""
    if not reminded_at or not completed_at:
        return "-"
    try:
        d = datetime.strptime(completed_at, "%Y-%m-%d %H:%M:%S") - \
            datetime.strptime(reminded_at, "%Y-%m-%d %H:%M:%S")
        total = int(d.total_seconds())
        if total < 0:
            return "-"
        m, s = divmod(total, 60)
        return f"{m}分{s}秒" if m else f"{s}秒"
    except ValueError:
        return "-"


def export_to_excel(records: list, dir_path: str, operator_name: str = "") -> str:
    """把记录写入 xlsx，返回生成的文件完整路径。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "完成记录"

    header_fill = PatternFill("solid", fgColor="1A8CFF")
    header_font = Font(color="FFFFFF", bold=True, name="微软雅黑")
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    for i, r in enumerate(records, start=1):
        ws.append([
            i,
            r.get("completed_at", ""),
            r.get("template_name", ""),
            r.get("reminder_content", ""),
            r.get("operator_name", ""),
            r.get("reminded_at", ""),
            r.get("delay_count", 0),
            _response_duration(r.get("reminded_at", ""), r.get("completed_at", "")),
        ])

    widths = [6, 20, 16, 40, 12, 20, 10, 12]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    os.makedirs(dir_path, exist_ok=True)
    safe_name = (operator_name or "全部").strip() or "全部"
    fname = f"提醒记录_{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(dir_path, fname)
    wb.save(path)
    return path
